import openpyxl
from dotenv import load_dotenv
from mysql.connector import connect, Error
import os
import requests
import csv

def MatchInfoExtractor(path, file, cursor, connection):
    matchInfo = {}
    fullFilePath = os.path.join(path, file)
    endMatchID = file.find('.')
    matchID = file[0:endMatchID]
    if (matchID.find("export") != -1):
        matchID = matchID[0: matchID.find("export")-1]

    query = f"""
        SELECT MatchID
        FROM Matches m
        WHERE m.MatchID = '{matchID}'
    """
    cursor.execute(query)

    queryResult = cursor.fetchall()

    if (len(queryResult) == 0):
        print(f"-----PROCESSING {matchID} -----------")
        matchInfo = getMatchInfo(fullFilePath)
        teamDict = getTeamDictionary(fullFilePath)
        if matchInfo != None:
            team1Score = 0
            team2Score = 0
            map = ""

            for key, value in matchInfo.items():
                if key == "Score team 1":
                    team1Score = int(value)
                
                if key == "Score team 2":
                    team2Score = int(value)

                if key == "Map":
                    map = value

            insertMatchQuery = f"""
                INSERT INTO Matches(MatchID, Team_1_Score, Team_2_Score, Map, MatchTime)
                VALUES
                ('{matchID}', {team1Score}, {team2Score}, '{map}', {0})
            """

            cursor.execute(insertMatchQuery)

            connection.commit()

            playerList = getPlayers(fullFilePath)

            for player in playerList:
                playerSearchQuery = f"""
                    SELECT SteamID
                    FROM Players p
                    WHERE p.SteamID = '{int(player)}'
                """
                cursor.execute(playerSearchQuery)
                queryResult = cursor.fetchall()

                if (len(queryResult) == 0):
                    insertPlayerQuery = f"""
                        INSERT INTO Players(SteamID, playerID)
                        VALUES
                        ('{int(player)}', {0})
                    """

                    cursor.execute(insertPlayerQuery)

                    connection.commit()
                else:
                    print(f"Player - {player} already added")

            if len(playerList) == 10:
                for i in range(1, 11):
                    playerInfo = getPlayersInfo(fullFilePath, i)

                    playerTeamNum = teamDict[playerInfo['Team']]

                    PlayerStatsInsertQuery = f"""
                        INSERT INTO PlayerStats (
                            steamID, MatchID, Team, Kills, Assists, Deaths,
                            Headshots, HeadshotsPerc, KR_Ratio, KD_Ratio,
                            TripleKills, QuadroKills, PentaKills, MVPs,
                            KAST, EntryKills, BombDefused, BombPlanted,
                            RWS, Rating, Rating2, ATD, ADR,
                            TradeKills, ClutchWinPer, totalSmokes,
                            totalFalshes, totalFire, totalHE
                        )
                        VALUES (
                            '{playerInfo["SteamID"]}', '{matchID}', {playerTeamNum}, {int(playerInfo["Kills"])}, {int(playerInfo["Assists"])}, {int(playerInfo["Deaths"])},
                            {int(playerInfo["HS"])}, {comma_to_float(playerInfo["HS%"])}, {comma_to_float(playerInfo["K/D"])}, {comma_to_float(playerInfo["KPR"])},
                            {int(playerInfo["3K"])}, {int(playerInfo["4K"])}, {int(playerInfo["5K"])}, {int(playerInfo["MVP"])},
                            {comma_to_float(playerInfo["KAST"])}, {int(playerInfo["Entry kill"])}, {int(playerInfo["Bomb defused"])}, {int(playerInfo["Bomb planted"])},
                            {comma_to_float(playerInfo["RWS"])}, {comma_to_float(playerInfo["Rating"])}, {comma_to_float(playerInfo["Rating 2"])}, {comma_to_float(playerInfo["ATD (s)"])}, {comma_to_float(playerInfo["ADR"])},
                            {int(playerInfo["Trade kill"])}, {0}, {int(playerInfo["Smoke"])},
                            {int(playerInfo["Flashbang"])}, {int(playerInfo["Incendiary"])}, {int(playerInfo["HE"])}
                        )
                    """

                    cursor.execute(PlayerStatsInsertQuery)

                    connection.commit()
            else:
                print(f"Not enough players for - {matchID}")
                
            print(f"Match - {matchID} - SUCCESSFUL")
        else:
            print(f"Match - {matchID} - ALREADY IN DB")


def comma_to_float(value):
    if isinstance(value, str):
        value = value.replace(',', '.')
    return float(value)


def getPlayersInfo(fullFilePath, playerNum):
    playerInfo = {}

    try:
        wb_obj = openpyxl.load_workbook(fullFilePath)

        worksheet = wb_obj['Players']

        for col in worksheet.iter_cols():
            key = col[0].value
            value = col[playerNum].value
            playerInfo[key] = value

        return playerInfo
    except Exception as e:
        print(e)
        return None
    
def getPlayers(fullFilePath):
    playerList = []

    try:
        wb_obj = openpyxl.load_workbook(fullFilePath)

        worksheet = wb_obj['Players']

        for row in worksheet.iter_rows():
            if row[1].value == 'SteamID':
                pass
            elif (row[1].value == None):
                break
            else:
                playerList.append(row[1].value)

        return playerList
    except Exception as e:
        print(e)
        return None
    
def getMatchInfo(fullFilePath):
    matchInfo = {}
    errors = 0 
    try:
        wb_obj = openpyxl.load_workbook(fullFilePath)

        worksheet = wb_obj['General']

        for col in worksheet.iter_cols():
            key = col[0].value
            value = col[1].value
            matchInfo[key] = value

        return matchInfo
    except Exception as e:
        print(e)
        return None

def getTeamDictionary(file):

    try:
        wb_obj = openpyxl.load_workbook(file)
        teamDict = {}
        sheet = wb_obj["General"]
        team1Alias = sheet.cell(row = 2, column = 13)
        team2Alias = sheet.cell(row= 2, column=14)

        teamDict[team1Alias.value] = 1
        teamDict[team2Alias.value] = 2

        return teamDict
    except Exception as e:
        return None

if __name__ == "__main__":
    load_dotenv()

    DB_USER = os.getenv("DB_USER")
    DB = os.getenv("CSGO_DB")
    DB_PW = os.getenv("DB_USER_PASSWORD")

    try:
        connection = connect(
            host = "localhost",
            user = DB_USER,
            password = DB_PW,
            database = DB
        )

        cursor = connection.cursor()
    except Error as e:
        print(e)

    ProcessedDemosPath = os.path.join(os.getcwd(), "ProcessedDemos")

    ProcessedDemos = os.listdir(ProcessedDemosPath)

    ProcessProgress = 0
    for file in ProcessedDemos:
        ProcessProgress += 1
        print(f"MATCH {ProcessProgress} OUT OF {len(ProcessedDemos)}")
        MatchInfoExtractor(ProcessedDemosPath, file, cursor, connection)